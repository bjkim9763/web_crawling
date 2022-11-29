from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import os
import openpyxl
from openpyxl.drawing.image import Image
import time
import urllib.request as req
from bs4 import BeautifulSoup

options = webdriver.ChromeOptions() # 크롬 옵션 객체 생성
options.add_argument('headless') # headless 모드 설정
options.add_argument("window-size=1920x1080") # 화면크기(전체화면)
options.add_argument("disable-gpu")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")
options.add_argument("user-agent=Mozilla/5.0 (Windows NT 6.1; WOW64; Trident/7.0; rv:11.0) like Gecko")
prefs = {'profile.default_content_setting_values': {'cookies' : 2, 'images': 2, 'plugins' : 2, 'popups': 2, 'geolocation': 2,
                                                    'notifications' : 2, 'auto_select_certificate': 2, 'fullscreen' : 2, 'mouselock' : 2,
                                                    'mixed_script': 2, 'media_stream' : 2, 'media_stream_mic' : 2, 'media_stream_camera': 2,
                                                    'protocol_handlers' : 2, 'ppapi_broker' : 2, 'automatic_downloads': 2, 'midi_sysex' : 2,
                                                    'push_messaging' : 2, 'ssl_cert_decisions': 2, 'metro_switch_to_desktop' : 2, 'protected_media_identifier': 2,
                                                    'app_banner': 2, 'site_engagement' : 2, 'durable_storage' : 2}}
options.add_experimental_option('prefs', prefs)

def get_blog_text(url):
    browser.get(url)
    time.sleep(1)
    try:
        pop_up = browser.find_element_by_css_selector('iframe#mainFrame')
        browser.switch_to.frame(pop_up)
        text = browser.find_element_by_css_selector('body').text
    except:
        text = browser.find_element_by_css_selector('body').text
    return text

search_word = input('검색할 단어 >>')

if not os.path.exists(f"./{search_word}.xlsx"):
    book = openpyxl.Workbook()
    book.save(f"./{search_word}.xlsx") # 엑셀

book = openpyxl.load_workbook(f"./{search_word}.xlsx")

sheet = book.active
sheet.column_dimensions['A'].width = "60"
sheet.column_dimensions['B'].width = "20"
sheet.column_dimensions['D'].width = "50"
sheet.cell(row=1,column=1).value = '제목'
sheet.cell(row=1,column=2).value = '스티커사진'
sheet.cell(row=1,column=3).value = '링크사진'
sheet.cell(row=1,column=4).value = '링크'
sheet.cell(row=1,column=5).value = 'fake?'
sheet.column_dimensions['C'].width = "20"
browser = webdriver.Chrome('./chromedriver.exe', options=options)
search = 'input#query'
browser.get('https://www.naver.com/')
time.sleep(1)
browser.find_element_by_css_selector(search).send_keys(search_word)
browser.find_element_by_css_selector(search).send_keys(Keys.ENTER)

list = browser.find_elements_by_css_selector('li.menu > a')
for i in list:
    if i.text == 'VIEW':
        i.click()
        break
time.sleep(1)
cnt = 0
cnt2 = 0
fake = {

}
if not os.path.exists("링크이미지"):
    os.mkdir("링크이미지")
if not os.path.exists("스티커이미지"):
    os.mkdir("스티커이미지")

browser.find_element_by_xpath('//*[@id="snb"]/div[1]/div/div[1]/a[2]').click()
for i in range(1000000000000000000000):
    # 사진 크롤링 맨 끝에 보통 원고료 표시 이미지로 fake확인
    if cnt % 58 == 0:
        cnt2 += 1
    for j in range(cnt2):
        browser.find_element_by_css_selector('html').send_keys(Keys.END)
        time.sleep(1)

    blog = browser.find_elements_by_css_selector('div.total_area > a')
    title = blog[cnt].text
    print(len(title), len(blog), cnt)
    url = blog[cnt].get_attribute('href')
    text = get_blog_text(url)

    img1 = browser.find_elements_by_css_selector('div.se-module.se-module-image img')
    img2 = browser.find_elements_by_css_selector('div.se-section.se-section-sticker.se-section-align-center.se-l-default img')
    if len(img1) != 0:
        try:
            url1 = img1[-1].get_attribute('src')
            req.urlretrieve(url1, f'링크이미지/{cnt + 2}.png')
            img_for_xl2 = Image(f'링크이미지/{cnt + 2}.png')
            sheet.add_image(img_for_xl2, f'B{cnt + 2}')
        except:
            print('')

    if len(img2) != 0:
        try:
            url2 = img2[-1].get_attribute('src')
            req.urlretrieve(url2, f'스티커이미지/{cnt + 2}.png')
            img_for_xl1 = Image(f'스티커이미지/{cnt + 2}.png')
            sheet.add_image(img_for_xl1, f'C{cnt + 2}')
        except:
            print('')

    if '원고료' in text:
        fake[title] = url
    elif '후원' in text:
        fake[title] = url
    elif '업체' in text:
        fake[title] = url
    elif '지원' in text:
        fake[title] = url
    elif '제공' in text:
        fake[title] = url
    # 글자로 적어놓은 fake여부 확인
    sheet.row_dimensions[cnt+2].height = 80
    cnt += 1
    sheet.cell(row=cnt+1, column=1).value = title
    sheet.cell(row=cnt+1, column=4).value = url
    if title in fake:
        sheet.cell(row=cnt+1, column=5).value = 1
    book.save(f"{search_word}.xlsx")
    browser.back()
    time.sleep(1)
    blog.clear()